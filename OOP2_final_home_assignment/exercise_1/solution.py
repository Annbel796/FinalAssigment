import csv
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.chart import BarChart, Reference
import os
from collections import defaultdict

class SalesDataProcessor:
    def __init__(self, input_file, output_file):
        self.input_file = input_file
        self.output_file = output_file
        self.workbook = Workbook()
        self.sales_sheet = self.workbook.active
        self.sales_sheet.title = "Sales Data"

    def load_csv_to_excel(self):
        with open(self.input_file, 'r', newline='', encoding='utf-8') as file:
            reader = csv.reader(file)
            headers = next(reader)
            headers.append("Total Sales")
            self.sales_sheet.append(headers)

            for cell in self.sales_sheet[1]:
                cell.font = Font(bold=True, size=14)

            for row in reader:
                count = int(row[3])
                price = float(row[4])
                total_sales = count * price
                row.append(total_sales)
                self.sales_sheet.append(row)

            for col in self.sales_sheet.columns:
                max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                adjusted_width = (max_length + 2)
                self.sales_sheet.column_dimensions[col[0].column_letter].width = adjusted_width

    def calculate_sales_by_region(self):
        ws_region = self.workbook.create_sheet(title="Sales by Region")
        ws_region.append(["Region", "Total Sales"])

        sales_by_region = defaultdict(float)

        for row in self.sales_sheet.iter_rows(min_row=2, values_only=True):
            region = row[2]
            total_sales = row[-1]
            sales_by_region[region] += total_sales

        for cell in ws_region[1]:
            cell.font = Font(bold=True, size=14)

        for region, total_sales in sales_by_region.items():
            ws_region.append([region, total_sales])

        for col in ws_region.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            adjusted_width = (max_length + 2)
            ws_region.column_dimensions[col[0].column_letter].width = adjusted_width

        self._create_chart(ws_region)

    def _create_chart(self, ws_region):
        chart = BarChart()
        chart.title = "Total Sales by Region"
        chart.x_axis.title = "Region"
        chart.y_axis.title = "Total Sales"
        chart.style = 10

        data = Reference(ws_region, min_col=2, min_row=1, max_row=ws_region.max_row)
        categories = Reference(ws_region, min_col=1, min_row=2, max_row=ws_region.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)

        ws_region.add_chart(chart, "D2")

    def save_excel(self):
        self.workbook.save(self.output_file)
        print(f"The Excel file '{self.output_file}' has been created and saved!")

if __name__ == "__main__":
    input_path = os.path.join(os.path.dirname(__file__), 'sales_data.csv')
    output_path = os.path.join(os.path.dirname(__file__), 'sales_output_v2.xlsx')

    if not os.path.exists(input_path):
        raise FileNotFoundError(f"Input file '{input_path}' not found.")
    
    processor = SalesDataProcessor(input_path, output_path)
    processor.load_csv_to_excel()
    processor.calculate_sales_by_region()
    processor.save_excel()
